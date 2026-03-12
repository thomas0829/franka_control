#!/usr/bin/env python3
"""
OpenPI-Polymetis Bridge
Connects OpenPI server (8000) to Franka robot via Polymetis (NUC)
Connects to existing robot services without re-launching them

Supported Models:
  - pi0_droid: Original PI0 model (action_horizon=10)
  - pi0_fast: Fast PI0 variant (action_horizon=10, same config as pi0_droid)
  - pi05_droid: PI05 model (action_horizon=8)

To switch models:
  1. Change MODEL_TYPE in Configuration section below
  2. Start corresponding OpenPI server:
     - PI0:      python scripts/serve_policy.py --policy.config=pi0_droid ...
     - PI0_FAST: python scripts/serve_policy.py --policy.config=pi0_droid ... (same as PI0)
     - PI05:     python scripts/serve_policy.py --policy.config=pi05_droid ...
"""
import time
import sys
import os
import signal
import atexit
import subprocess
import numpy as np
import cv2
from datetime import datetime
from pathlib import Path

# Excel logging imports (only used if ENABLE_EXCEL_LOGGING is True)
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# Global flag for emergency stop
_emergency_stop = False

def signal_handler(signum, frame):
    """Handle Ctrl+C by setting emergency stop flag"""
    global _emergency_stop
    _emergency_stop = True
    print("\n\n🛑 Ctrl+C detected - Emergency stop triggered!", flush=True)

# Register signal handler
signal.signal(signal.SIGINT, signal_handler)

# Add project path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
OPENPI_CLIENT_DIR = os.path.abspath(
    os.path.join(os.path.dirname(__file__), '..', '..', 'openpi', 'packages', 'openpi-client')
)

# Import OpenPI client
try:
    from openpi_client.websocket_client_policy import WebsocketClientPolicy
except ImportError:
    print("Error: openpi_client not found in this Python environment:")
    print(f"  Python: {sys.executable}")
    if os.path.isdir(OPENPI_CLIENT_DIR):
        print("  Install with one of:")
        print(f"    {sys.executable} -m pip install -e {OPENPI_CLIENT_DIR}")
        print(f"    uv pip install -e {OPENPI_CLIENT_DIR} --python {sys.executable}")
    else:
        print("  Install openpi-client into the same Python interpreter used to run this script.")
    sys.exit(1)

# Import ZED SDK
try:
    import pyzed.sl as sl
    ZED_AVAILABLE = True
except ImportError:
    print("Warning: ZED SDK not found. ZED camera will not be available.")
    ZED_AVAILABLE = False

# ============ Configuration ============
# Model Selection: pi05_droid or pi0_droid or pi0_fast"
MODEL_TYPE = "pi05_droid"

# ============ Auto Server Management ============
# Set to True to automatically start/stop OpenPI server based on MODEL_TYPE
AUTO_START_SERVER = True  # True: Auto manage server, False: Manual (like before)

# Model server configurations - maps MODEL_TYPE to server startup command
# Compute openpi directory relative to this script
OPENPI_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'openpi'))

MODEL_SERVER_CONFIGS = {
    "pi05_droid": {
        "dir": OPENPI_DIR,
        "cmd": "uv run scripts/serve_policy.py policy:checkpoint --policy.config=pi05_droid --policy.dir=gs://openpi-assets/checkpoints/pi05_droid",
        "display_name": "Pi0.5-velocity"
    },
    "pi0_droid": {
        "dir": OPENPI_DIR,
        "cmd": "uv run scripts/serve_policy.py policy:checkpoint --policy.config=pi0_droid --policy.dir=gs://openpi-assets/checkpoints/pi0_droid",
        "display_name": "Pi0-velocity"
    },
    "pi0_fast": {
        "dir": OPENPI_DIR,
        "cmd": "uv run scripts/serve_policy.py policy:checkpoint --policy.config=pi0_fast_droid --policy.dir=gs://openpi-assets/checkpoints/pi0_fast_droid",
        "display_name": "Pi0-fast-velocity"
    }
}

# Camera Configuration
# System: 2x ZED cameras (official configuration)
# - ZED 2 for external view (shoulder camera)
# - ZED Mini for wrist-mounted view
USE_ZED_FOR_BOTH = True             # Use ZED for both cameras (official config)
USE_ZED_AS_WRIST = True             # ZED Mini on wrist
USE_REALSENSE_WRIST = False         # No RealSense cameras

NUC_IP = "192.168.1.6"              # Polymetis zerorpc server IP (port 4242)
OPENPI_HOST = "127.0.0.1"           # OpenPI server IP
OPENPI_PORT = 8000                  # OpenPI server port
PROMPT = "Put the doll on the plate."  # Task instruction for OpenPI
# ============ Loop Recording Configuration ============
# Set LOOP to the number of episodes you want to record
# LOOP = 0: Stop record
# LOOP = 1: Record 1 episode (single recording)
# LOOP = 3: Record 3 episodes (will loop 3 times)
# Each episode will be saved in a separate timestamped folder under a task-named directory
LOOP = 0  # Number of episodes to record per position per model

# ============ Position Configuration ============
# New recording scheme:
# - 6 models (3 velocity: pi0, pi0_fast, pi05 + 3 position: pi0, pi0_fast, pi05)
# - Each model runs 9 times: 3 initial positions × 3 videos per position
# - Total: 6 models × 9 runs = 54 videos per task
ENABLE_POSITION_VARIANT = False  # Set to True to use position variants and log to Excel
POSITION_VARIANT = "pos-1"  # Current position: "pos-1", "pos-2", or "pos-3" (only used if ENABLE_POSITION_VARIANT is True)

# ============ Excel Logging Configuration ============
# Set to True to enable Excel logging of episode metadata
# When enabled, you will be prompted after each episode for:
# - Model name (e.g., "pi0", "pi05")
# - Success status (yes/no)
# Data logged: Task, Episode, Model, Success, Video Path, Timestamp
# Excel file will be saved in the vid directory as "episode_log.xlsx"
ENABLE_EXCEL_LOGGING = False  # True: Enable logging, False: Disable

# ZED Camera Configuration (Official OpenPI DROID Setup)
# Both cameras now working on USB 3.0
# Using Serial Numbers instead of IDs for stability (IDs can change after reconnection)
# If a camera was replaced, re-check current serial numbers with:
#   python - <<'PY'
#   import pyzed.sl as sl
#   for d in sl.Camera.get_device_list():
#       print(d.camera_model, d.serial_number, d.camera_state, d.path)
#   PY
# If a camera shows up as NOT AVAILABLE, inspect USB-level serials with:
#   for dev in $(lsusb -d 2b03: | awk '{print $6}'); do
#       echo "== $dev =="; lsusb -v -d "$dev" 2>/dev/null | rg 'iProduct|iSerial'
#   done
ZED_EXTERNAL_ID = None              # Not using ID, using SN instead
ZED_EXTERNAL_SN = 26706125          # ZED 2 (external/shoulder view)
ZED_WRIST_ID = None                 # Not using ID, using SN instead
ZED_WRIST_SN = 15679333             # ZED Mini (wrist view)
WIDTH, HEIGHT, FPS = 1280, 720, 15  # ZED HD720 mode @ 15fps (official DROID config)
# Note: Using SN ensures cameras are always correctly identified even after USB reconnection

DROID_CONTROL_FREQUENCY = 15  # Hz

# Maximum steps before stopping
MAX_STEPS = 10000

# Action horizon (official OpenPI settings):
# PI0/PI0_FAST: action_horizon=10 (official training), PI05: action_horizon=16 
if MODEL_TYPE in ["pi0_droid", "pi0_fast"]:
    OPEN_LOOP_HORIZON = 10  # Official PI0 setting - matches training horizon
else:
    OPEN_LOOP_HORIZON = 8   # Works well for PI05
                        # Note: Each action's joint_delta is computed from current robot_state
                        # This is correct behavior - velocity actions should be relative to current state

# SAFETY: Velocity scaling factor (for cautious testing)
VELOCITY_SCALE = 1.0  # Full speed - using safety limits instead of velocity scaling

# Derived parameters
CTRL_HZ = DROID_CONTROL_FREQUENCY

# ============ Auto Server Management Functions ============
_server_process = None  # Global variable to track server process

def start_openpi_server(model_type):
    """Start OpenPI server for the specified model type"""
    global _server_process
    
    if model_type not in MODEL_SERVER_CONFIGS:
        raise ValueError(f"Unknown model type: {model_type}. Available: {list(MODEL_SERVER_CONFIGS.keys())}")
    
    config = MODEL_SERVER_CONFIGS[model_type]
    display_name = config.get('display_name', model_type)
    
    print(f"\n{'='*60}")
    print(f"🚀 Starting OpenPI server")
    print(f"{'='*60}")
    print(f"Model: {display_name} ({model_type})")
    print(f"Directory: {config['dir']}")
    print(f"Command: {config['cmd']}")
    print(f"Port: {OPENPI_PORT}")
    print()
    
    # Start server process
    _server_process = subprocess.Popen(
        config['cmd'],
        shell=True,
        cwd=config['dir'],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        preexec_fn=os.setsid  # Create new process group for easy cleanup
    )
    
    print(f"Server process started (PID: {_server_process.pid})")
    print(f"Waiting for {display_name} server to be ready", end="", flush=True)
    
    # Wait for server to be ready (check if port is listening)
    import socket
    max_wait = 120  # Wait up to 2 minutes
    for i in range(max_wait):
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(1)
            result = sock.connect_ex(('127.0.0.1', OPENPI_PORT))
            sock.close()
            if result == 0:
                print(" ✓")
                print(f"{display_name} server is ready on port {OPENPI_PORT}!")
                time.sleep(3)  # Extra wait to ensure fully initialized
                return True
        except:
            pass
        print(".", end="", flush=True)
        time.sleep(1)
    
    print(" ✗")
    print(f"⚠️  Warning: Server did not respond within {max_wait} seconds")
    return False

def stop_openpi_server():
    """Stop the currently running OpenPI server"""
    global _server_process
    
    if _server_process is None:
        return
    
    print(f"\n🛑 Stopping OpenPI server (PID: {_server_process.pid})...")
    
    try:
        # Kill entire process group (includes all child processes)
        os.killpg(os.getpgid(_server_process.pid), signal.SIGTERM)
        _server_process.wait(timeout=10)
        print("Server stopped successfully")
    except subprocess.TimeoutExpired:
        print("Server didn't stop gracefully, forcing...")
        try:
            os.killpg(os.getpgid(_server_process.pid), signal.SIGKILL)
        except:
            pass
    except Exception as e:
        print(f"Error stopping server: {e}")
    
    _server_process = None
    time.sleep(2)  # Wait for port to be released

def cleanup_server():
    """Cleanup function to ensure server is stopped on exit"""
    stop_openpi_server()

# Register cleanup function
atexit.register(cleanup_server)

print(f"Control frequency: {CTRL_HZ} Hz (official DROID)")
print(f"Policy query frequency: ~{CTRL_HZ / OPEN_LOOP_HORIZON:.2f} Hz")
print(f"Open loop horizon: {OPEN_LOOP_HORIZON} steps")
print(f"Velocity control mode: TRUE_VELOCITY (direct velocity commands)")
print(f"Velocity scale: {VELOCITY_SCALE:.2f} (safety factor)")

def print_zed_sn_lookup_help():
    """Print commands for checking current ZED serial numbers."""
    print("  To check current ZED serial numbers:")
    print("    python - <<'PY'")
    print("    import pyzed.sl as sl")
    print("    for d in sl.Camera.get_device_list():")
    print("        print(d.camera_model, d.serial_number, d.camera_state, d.path)")
    print("    PY")
    print("  If a camera shows NOT AVAILABLE, check USB-level serials:")
    print("    for dev in $(lsusb -d 2b03: | awk '{print $6}'); do")
    print("        echo \"== $dev ==\"; lsusb -v -d \"$dev\" 2>/dev/null | rg 'iProduct|iSerial'")
    print("    done")

def print_zed_device_list():
    """Print the current device list seen by the ZED SDK."""
    if not ZED_AVAILABLE:
        return
    try:
        devices = sl.Camera.get_device_list()
    except Exception as exc:
        print(f"  Could not query ZED device list: {exc}")
        return
    print("  ZED SDK device list:")
    if not devices:
        print("    <empty>")
        return
    for i, dev in enumerate(devices):
        print(f"    [{i}] model={dev.camera_model} sn={dev.serial_number} state={dev.camera_state} path={dev.path}")

def open_zed_camera(camera_id=None, serial_number=None, width=1280, height=720, fps=30):
    """Open ZED camera with optimized settings
    
    Args:
        camera_id: Camera ID (0 for first camera, 1 for second, etc.) - use this OR serial_number
        serial_number: Camera serial number - use this if camera_id doesn't work
        width, height: Resolution (1280x720 for HD720)
        fps: Frame rate (15 or 30)
    """
    if not ZED_AVAILABLE:
        raise RuntimeError("ZED SDK is not available. Please install pyzed.")
    
    if camera_id is None and serial_number is None:
        raise ValueError("Must provide either camera_id or serial_number")
    
    zed = sl.Camera()
    
    # Set initialization parameters
    init_params = sl.InitParameters()
    init_params.camera_resolution = sl.RESOLUTION.HD720  # 1280x720
    init_params.camera_fps = fps
    init_params.depth_mode = sl.DEPTH_MODE.NONE  # We only need RGB, no depth
    init_params.coordinate_units = sl.UNIT.METER
    
    # Set camera ID or serial number to select specific camera
    if camera_id is not None:
        init_params.set_from_camera_id(camera_id)
        identifier = f"ID {camera_id}"
    else:
        init_params.set_from_serial_number(serial_number)
        identifier = f"SN {serial_number}"
    
    # Open the camera
    err = zed.open(init_params)
    if err != sl.ERROR_CODE.SUCCESS:
        raise RuntimeError(f"Failed to open ZED camera {identifier}: {err}")
    
    # Get camera information
    cam_info = zed.get_camera_information()
    camera_model = cam_info.camera_model
    actual_serial_number = cam_info.serial_number
    
    # Set camera controls for better image quality
    # Auto exposure and white balance (default, works well in most cases)
    zed.set_camera_settings(sl.VIDEO_SETTINGS.EXPOSURE, -1)  # -1 = auto
    zed.set_camera_settings(sl.VIDEO_SETTINGS.GAIN, -1)      # -1 = auto
    zed.set_camera_settings(sl.VIDEO_SETTINGS.WHITEBALANCE_AUTO, 1)  # Enable auto white balance
    
    # Set runtime parameters (minimal for RGB-only mode)
    runtime_params = sl.RuntimeParameters()
    
    return zed, runtime_params, camera_model, actual_serial_number


def get_zed_image(zed, runtime_params, rotate_180=False):
    """Get RGB image from ZED camera
    
    Args:
        zed: ZED camera object
        runtime_params: ZED runtime parameters
        rotate_180: If True, rotate image 180 degrees (for upside-down mounted cameras)
    """
    image = sl.Mat()
    # Grab a new frame
    if zed.grab(runtime_params) == sl.ERROR_CODE.SUCCESS:
        # Retrieve left RGB image
        zed.retrieve_image(image, sl.VIEW.LEFT)
        # Convert to numpy array (BGRA format)
        img_bgra = image.get_data()
        # Convert BGRA to BGR
        img_bgr = cv2.cvtColor(img_bgra, cv2.COLOR_BGRA2BGR)
        # Rotate 180 degrees if needed (e.g., wrist camera mounted upside down)
        if rotate_180:
            img_bgr = cv2.rotate(img_bgr, cv2.ROTATE_180)
        return img_bgr
    else:
        return None


# Global camera references for cleanup
_zed_ext = None
_zed_wri = None
_robot_ref = None  # Global reference to robot for emergency cleanup

def cleanup_cameras():
    """Cleanup function to ensure cameras are properly closed"""
    global _zed_ext, _zed_wri
    
    print("\nCleaning up cameras...")
    try:
        if _zed_ext is not None:
            _zed_ext.close()
            print("  ✓ External ZED camera closed")
            _zed_ext = None
    except Exception as e:
        print(f"  ⚠ Error closing external camera: {e}")
    
    try:
        if _zed_wri is not None:
            _zed_wri.close()
            print("  ✓ Wrist ZED camera closed")
            _zed_wri = None
    except Exception as e:
        print(f"  ⚠ Error closing wrist camera: {e}")
    
    # Add a small delay to ensure cleanup completes
    time.sleep(0.5)

def emergency_stop_robot():
    """Emergency stop - just stop the robot, don't move anywhere"""
    global _robot_ref
    
    if _robot_ref is None:
        return
    
    try:
        print("\n🛑 Emergency stop - Stopping robot...")
        
        # Just stop any motion - don't try to move anywhere
        zero_velocity = [0.0] * 7 + [0.0]
        _robot_ref.update_command(
            command=zero_velocity,
            action_space="joint_velocity",
            gripper_action_space="position",
            blocking=False
        )
        time.sleep(0.2)
        print("   ✓ Robot stopped")
    except Exception as e:
        print(f"   ⚠ Error stopping robot: {e}")

def reset_robot_to_home():
    """Reset robot to home position safely (only for normal completion)"""
    global _robot_ref
    
    if _robot_ref is None:
        return
    
    try:
        print("\n🔄 Returning robot to home position...")
        
        # Use the same home position as everywhere else in the code
        home_joints = np.array([0.0, -0.5, 0.0, -2.40, 0.0, 1.90, 0.0])
        # Use blocking=True for slower, safer movement
        _robot_ref.update_joints(command=home_joints.tolist(), velocity=False, blocking=True)
        _robot_ref.update_gripper(command=0.0, velocity=False, blocking=True)
        print("   ✓ Robot returned to home position")
    except Exception as e:
        print(f"   ⚠ Error resetting robot: {e}")

def signal_handler(sig, frame):
    """Handle termination signals"""
    global _emergency_stop
    
    print(f"\n\n🛑 SIGNAL RECEIVED: {sig}")
    print("   Setting emergency stop flag...")
    
    # Set emergency stop flag
    _emergency_stop = True
    
    # Raise KeyboardInterrupt to interrupt blocking operations like input()
    # This is safe now because we handle it gracefully in try-except blocks
    raise KeyboardInterrupt("User requested emergency stop")

def cleanup_all():
    """Complete cleanup: robot + cameras (for atexit)"""
    # Don't reset robot in atexit - could be dangerous if program crashed
    # Just cleanup cameras
    cleanup_cameras()

# Register cleanup handlers
atexit.register(cleanup_all)
signal.signal(signal.SIGINT, signal_handler)   # Ctrl+C
signal.signal(signal.SIGTERM, signal_handler)  # kill command


def bgr_to_rgb(img):
    """Convert BGR to RGB"""
    return cv2.cvtColor(img, cv2.COLOR_BGR2RGB)


def resize_with_pad(img, height=224, width=224):
    """Resize image with padding to maintain aspect ratio (like OpenPI official).
    
    This mimics the behavior of openpi_client.image_tools.resize_with_pad.
    Reduces network transfer by ~33x (1280x720 -> 224x224).
    """
    from PIL import Image
    
    # Convert to PIL Image
    pil_img = Image.fromarray(img)
    
    cur_width, cur_height = pil_img.size
    if cur_width == width and cur_height == height:
        return img
    
    # Calculate resize ratio
    ratio = max(cur_width / width, cur_height / height)
    resized_height = int(cur_height / ratio)
    resized_width = int(cur_width / ratio)
    
    # Resize image
    resized_image = pil_img.resize((resized_width, resized_height), resample=Image.BILINEAR)
    
    # Create zero-padded canvas
    zero_image = Image.new(resized_image.mode, (width, height), 0)
    pad_height = max(0, int((height - resized_height) / 2))
    pad_width = max(0, int((width - resized_width) / 2))
    zero_image.paste(resized_image, (pad_width, pad_height))
    
    return np.array(zero_image)


def main():
    global _emergency_stop, _robot_ref  # Declare at start of main function
    
    # Get display name for model
    model_display_name = MODEL_SERVER_CONFIGS.get(MODEL_TYPE, {}).get('display_name', MODEL_TYPE)
    
    print("=" * 60)
    print("OpenPI-Polymetis Bridge")
    print("=" * 60)
    print(f"Model: {model_display_name} ({MODEL_TYPE})")
    print(f"Auto start server: {AUTO_START_SERVER}")
    print(f"NUC IP: {NUC_IP}:4242 (zerorpc)")
    print(f"OpenPI: {OPENPI_HOST}:{OPENPI_PORT}")
    print(f"Task: {PROMPT}")
    print(f"🔁 Episodes to record: {LOOP}")
    print("=" * 60)
    
    # Auto start OpenPI server if enabled
    if AUTO_START_SERVER:
        print(f"\n[0/4] Auto-starting OpenPI server for {MODEL_TYPE}...")
        if not start_openpi_server(MODEL_TYPE):
            print("⚠️  Server startup failed or timed out")
            print("You can try:")
            print("  1. Set AUTO_START_SERVER = False and start server manually")
            print("  2. Check if model directory and command are correct")
            return
    
    # Connect to OpenPI server
    print("\n[1/4] Connecting to OpenPI server...")
    try:
        openpi_client = WebsocketClientPolicy(host=OPENPI_HOST, port=OPENPI_PORT)
        server_metadata = openpi_client.get_server_metadata()
        print(f"✓ OpenPI connected")
        print(f"  Server metadata: {server_metadata}")
    except Exception as e:
        print(f"✗ OpenPI connection failed: {e}")
        if AUTO_START_SERVER:
            print("Server was auto-started but connection failed.")
        else:
            print("Make sure OpenPI server is running on port 8000")
        return
    
    # Connect to robot (same as GELLO: launch=True will auto-start robot on NUC)
    print("\n[2/4] Connecting to robot on NUC...")
    try:
        from robot.real.server_interface import ServerInterface
        
        # Connect to zerorpc and auto-launch robot (same as GELLO)
        print(f"  Connecting to zerorpc at {NUC_IP}:4242...")
        print("  (This will auto-launch Polymetis robot_server on NUC)")
        robot = ServerInterface(ip_address=NUC_IP)  # launch=True by default!
        _robot_ref = robot  # Store global reference for emergency cleanup
        print("  ✓ Robot launched and connected")
        
        # Test connection by getting robot state
        print("  Testing robot state...")
        joint_pos = robot.get_joint_positions()
        print(f"  ✓ Got joint positions: {joint_pos[:3]}... (showing first 3)")
        
        # Reset to home position at startup
        print("\n  Resetting robot to home position...")
        # Standard home position
        home_joints = np.array([
            0.0, -0.5, 0.0, -2.40, 
            0.0, 1.90, 0.0
        ])
        # J1 = -0.5 rad (-29°): Standard home height
        
        current_joints = robot.get_joint_positions()
        print(f"    Current: {np.round(current_joints, 3).tolist()}")
        print(f"    Target:  {np.round(home_joints, 3).tolist()}")
        
        # Move to home position (blocking=True for slower movement)
        robot.update_joints(
            command=home_joints.tolist(),
            velocity=False,
            blocking=True
        )
        
        # Open gripper (0.0 = open for Robotiq)
        robot.update_gripper(command=0.0, velocity=False, blocking=True)
        print("  ✓ Robot reset to home position")
        
        print("✓ Robot interface ready")
        
    except Exception as e:
        print(f"✗ Robot connection failed: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Open cameras
    print(f"\n[3/4] Initializing cameras...")
    print("Camera setup: 2x ZED cameras (official OpenPI configuration)")
    print(f"  - ZED 2 (external/shoulder view): SN {ZED_EXTERNAL_SN}")
    print(f"  - ZED Mini (wrist view): SN {ZED_WRIST_SN}")
    
    # Initialize variables
    global _zed_ext, _zed_wri  # Global references for cleanup
    ext_p = None  # Not used (RealSense external)
    wri_p = None  # Not used (RealSense wrist)
    zed_ext = None
    zed_ext_runtime = None
    zed_wri = None
    zed_wri_runtime = None
    
    # Camera initialization with auto-retry
    MAX_RETRIES = 3
    RETRY_DELAY = 2.0  # seconds
    
    # Initialize external ZED camera (ZED 2) with retry
    for attempt in range(MAX_RETRIES):
        try:
            if attempt > 0:
                print(f"\n  Retry {attempt}/{MAX_RETRIES-1} for external camera...")
                time.sleep(RETRY_DELAY)
            
            zed_ext, zed_ext_runtime, ext_model, ext_sn = open_zed_camera(
                serial_number=ZED_EXTERNAL_SN, 
                width=WIDTH, 
                height=HEIGHT, 
                fps=FPS
            )
            _zed_ext = zed_ext  # Store for cleanup
            print(f"✓ External ZED camera (SN: {ext_sn})")
            print(f"  Model: {ext_model}")
            # Warm up camera
            for _ in range(10):
                get_zed_image(zed_ext, zed_ext_runtime)
            break  # Success, exit retry loop
        except Exception as e:
            print(f"✗ External ZED camera initialization failed (attempt {attempt+1}/{MAX_RETRIES}): {e}")
            if attempt == MAX_RETRIES - 1:
                print("✗ External camera failed after all retries")
                print_zed_device_list()
                print_zed_sn_lookup_help()
                cleanup_cameras()
                return
    
    # Initialize wrist ZED camera (ZED Mini) with retry
    for attempt in range(MAX_RETRIES):
        try:
            if attempt > 0:
                print(f"\n  Retry {attempt}/{MAX_RETRIES-1} for wrist camera...")
                time.sleep(RETRY_DELAY)
            
            zed_wri, zed_wri_runtime, wri_model, wri_sn = open_zed_camera(
                serial_number=ZED_WRIST_SN, 
                width=WIDTH, 
                height=HEIGHT, 
                fps=FPS
            )
            _zed_wri = zed_wri  # Store for cleanup
            print(f"✓ Wrist ZED camera (SN: {wri_sn})")
            print(f"  Model: {wri_model}")
            # Warm up camera
            for _ in range(10):
                get_zed_image(zed_wri, zed_wri_runtime)  # No rotation
            break  # Success, exit retry loop
        except Exception as e:
            print(f"✗ Wrist ZED camera initialization failed (attempt {attempt+1}/{MAX_RETRIES}): {e}")
            if attempt == MAX_RETRIES - 1:
                print("✗ Wrist camera failed after all retries")
                print_zed_device_list()
                print_zed_sn_lookup_help()
                cleanup_cameras()
                return
    
    # Save initial camera views for debugging
    print("\n📸 Capturing camera views at reset position...")
    try:
        # Get External RealSense image
        ext_img_debug = None
        wri_img_debug = None
        
        # Get External ZED image
        if zed_ext:
            ext_img_debug = get_zed_image(zed_ext, zed_ext_runtime)
        
        # Get Wrist ZED image
        if zed_wri:
            wri_img_debug = get_zed_image(zed_wri, zed_wri_runtime)  # No rotation
        
        # Save combined image (one file, overwritten each time)
        if ext_img_debug is not None and wri_img_debug is not None:
            # Resize to same height for horizontal stacking
            h = 480
            ext_resized = cv2.resize(ext_img_debug, (int(ext_img_debug.shape[1] * h / ext_img_debug.shape[0]), h))
            wri_resized = cv2.resize(wri_img_debug, (int(wri_img_debug.shape[1] * h / wri_img_debug.shape[0]), h))
            combined = np.hstack([ext_resized, wri_resized])
            
            # Add labels
            cv2.putText(combined, "External (ZED 2)", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
            cv2.putText(combined, "Wrist (ZED Mini)", (ext_resized.shape[1] + 20, 40), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
            
            debug_path = Path(__file__).parent / "debug_cameras.jpg"
            cv2.imwrite(str(debug_path), combined)
            print("  ✓ Saved: debug_cameras.jpg (combined view)\n")
        
    except Exception as e:
        print(f"  ⚠️  Failed to save debug image: {e}\n")
    
    # Preview cameras and wait for confirmation
    print("\n[4/5] Camera Preview")
    print("=" * 60)
    print("Displaying camera feeds...")
    print("Check the camera window to verify:")
    print("  - External ZED 2 (left) shows the scene correctly")
    print("  - Wrist ZED Mini (right) shows the robot gripper")
    print("  - Robot is at home position")
    print("  - Scene is set up correctly")
    print("=" * 60)
    print("\n👁️  Press ENTER when ready to start inference, or Ctrl+C to quit")
    
    try:
        # Preview loop - show cameras until user presses Enter
        preview_counter = 0
        while True:
            # Collect active camera images
            cameras_to_show = []
            camera_labels = []
            
            # Get External ZED image
            ext_img = None
            if zed_ext:
                ext_img = get_zed_image(zed_ext, zed_ext_runtime)
                if ext_img is not None:
                    cameras_to_show.append(ext_img)
                    camera_labels.append("External (ZED 2)")
            
            # Get Wrist ZED image
            wri_img = None
            if zed_wri:
                wri_img = get_zed_image(zed_wri, zed_wri_runtime)  # No rotation
                if wri_img is not None:
                    cameras_to_show.append(wri_img)
                    camera_labels.append("Wrist (ZED Mini)")
            
            # Skip if no wrist image available
            if wri_img is None or ext_img is None:
                continue
            
            # Only update display every 3 frames for smooth preview
            if preview_counter % 3 == 0 and len(cameras_to_show) > 0:
                # Resize all cameras (doubled size: 848x480 instead of 424x240)
                displays = [cv2.resize(img, (848, 480)) for img in cameras_to_show]
                
                # Stack horizontally
                combined = np.hstack(displays)
                
                # Add text labels (larger font for bigger window)
                for i, label in enumerate(camera_labels):
                    x_pos = 20 + i * 860
                    cv2.putText(combined, label, (x_pos, 50), 
                               cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 255, 0), 3)
                
                # Add instruction text (centered based on number of cameras)
                text_x = 280 if len(cameras_to_show) == 2 else 600
                cv2.putText(combined, "Press ENTER to start", (text_x, 450), 
                           cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 255, 255), 2)
                
                cv2.imshow("Camera Preview - Press ENTER to start", combined)
            
            preview_counter += 1
            
            # Check for Enter key or window close (check every frame for responsiveness)
            key = cv2.waitKey(1)  # 1ms instead of 30ms for better responsiveness
            if key == 13:  # Enter key
                print("\n✓ Starting inference...")
                cv2.destroyAllWindows()  # Close the preview window
                break
            elif key == 27:  # ESC key
                print("\n✗ Cancelled by user")
                cv2.destroyAllWindows()
                cleanup_cameras()
                return
                
    except KeyboardInterrupt:
        print("\n✗ Cancelled by user")
        cv2.destroyAllWindows()
        cleanup_cameras()
        return
    
    dt = 1.0 / CTRL_HZ
    print(f"\n[5/5] Starting control loop")
    print(f"Control Frequency: {CTRL_HZ} Hz (DROID official)")
    print(f"Policy Query Frequency: ~{CTRL_HZ / OPEN_LOOP_HORIZON:.1f} Hz (every {OPEN_LOOP_HORIZON} steps)")
    print("\n⌨️  Controls:")
    print("  - Press 'Q' to stop current episode early (saves and continues to next)")
    print("  - Press Ctrl+C to emergency stop (saves and exits program)")
    print("-" * 60)
    
    # Gripper state tracking for hysteresis (prevent oscillation)
    last_gripper_cmd = None
    
    # ===== Setup base directory for loop recording =====
    # Use relative path based on script location
    script_dir = Path(__file__).parent
    vid_base_dir = script_dir / "vid"
    vid_base_dir.mkdir(exist_ok=True, parents=True)
    
    # Create task-based folder with position for all recording modes
    if LOOP > 0:
        # Convert task to folder name (e.g., "Pick up the cube" -> "pick_up_the_cube")
        task_folder_name = PROMPT.lower().replace(" ", "_").replace(".", "")
        # Structure: vid/task_name/pos-Y/model_name/ (if ENABLE_POSITION_VARIANT) or vid/task_name/model_name/
        if ENABLE_POSITION_VARIANT:
            pos_dir = vid_base_dir / task_folder_name / POSITION_VARIANT
            model_dir = pos_dir / MODEL_TYPE
        else:
            model_dir = vid_base_dir / task_folder_name / MODEL_TYPE
        task_dir = model_dir  # Final directory for this session
        task_dir.mkdir(parents=True, exist_ok=True)
        if LOOP > 1:
            print(f"\n🔁 Recording {LOOP} episodes")
        else:
            print(f"\n🔁 Recording 1 episode")
        if ENABLE_POSITION_VARIANT:
            print(f"   Position: {POSITION_VARIANT}")
        print(f"   Model: {MODEL_TYPE}")
        print(f"   Save to: {task_dir}")
    else:
        print(f"\n♾️  Infinite inference mode (LOOP=0)")
        print(f"   Press Ctrl+C to stop anytime")
        print(f"   No video recording")
    
    # Determine number of iterations
    # LOOP=0 means infinite loop (for testing/demos without recording)
    infinite_mode = (LOOP == 0)
    num_iterations = float('inf') if infinite_mode else LOOP
    
    # Flag to track if we should exit (Ctrl+C or error)
    should_exit = False
    
    # ===== Main recording loop (iterate for each episode) =====
    completed_episodes = 0  # Track number of successfully completed episodes
    current_episode = 1     # Current episode number being recorded
    session_model_name = None  # Store model name for entire session (ask once)
    
    while completed_episodes < num_iterations:
        # Check emergency stop at start of each episode
        if _emergency_stop:
            print("\n🛑 Emergency stop - Exiting episode loop", flush=True)
            break
        
        if LOOP > 1:
            print(f"\n{'='*70}")
            print(f"📹 EPISODE {current_episode} (Completed: {completed_episodes}/{LOOP})")
            print(f"{'='*70}")
        
        step = 0
        last_print_time = time.time()
        
        # Episode-specific preview image (will be captured from first frame)
        episode_preview_image = None
        
        # Action chunking variables (OpenPI official approach)
        pred_action_chunk = None
        actions_from_chunk_completed = 0
        
        # Trajectory recording for debugging
        trajectory_log = []
        
        # ===== Setup video recording for this episode (skip if LOOP=0) =====
        if not infinite_mode:
            # Create episode folder with timestamp
            session_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            session_dir = task_dir / session_timestamp
            session_dir.mkdir(exist_ok=True)
            
            # Save prompt to text file
            instruction_file = session_dir / "instruction.txt"
            with open(instruction_file, "w") as f:
                f.write(f"Session: {session_timestamp}\n")
                f.write(f"Model: {MODEL_TYPE}\n")
                f.write(f"Instruction: {PROMPT}\n")
                if LOOP >= 1:
                    f.write(f"Episode: {current_episode} (Progress: {completed_episodes}/{LOOP} completed)\n")
                f.write(f"Steps: TBD\n")  # Will be updated at the end
                f.write(f"Control Frequency: {DROID_CONTROL_FREQUENCY} Hz\n")
                f.write(f"Success: TBD\n")  # Will be updated at the end
            
            print(f"\n📹 Video Recording Setup:")
            if LOOP >= 1:
                print(f"   Episode: {current_episode} (Progress: {completed_episodes}/{LOOP} completed)")
            print(f"   Session: {session_timestamp}")
            print(f"   Directory: {session_dir}")
            print(f"   Instruction saved to: {instruction_file}")
            
            # Initialize video writers with mp4v (works reliably)
            fourcc = cv2.VideoWriter_fourcc(*'mp4v')
            print(f"   Using codec: mp4v (MPEG-4)")
            
            shoulder_video = cv2.VideoWriter(
                str(session_dir / "shoulder_view.mp4"),
                fourcc, DROID_CONTROL_FREQUENCY, (WIDTH, HEIGHT)
            )
            wrist_video = cv2.VideoWriter(
                str(session_dir / "wrist_view.mp4"),
                fourcc, DROID_CONTROL_FREQUENCY, (WIDTH, HEIGHT)
            )
            
            # Verify writers are opened successfully
            if not shoulder_video.isOpened() or not wrist_video.isOpened():
                print(f"   ⚠ ERROR: Failed to open video writers!")
                print(f"   Codec: mp4v")
                print(f"   This may cause video recording to fail")
            
            print(f"   Shoulder video: {session_dir / 'shoulder_view.mp4'}")
            print(f"   Wrist video: {session_dir / 'wrist_view.mp4'}")
            print(f"   Recording will stop after {MAX_STEPS} steps\n")
        else:
            print(f"\n🚀 Starting inference (no recording)")
            print(f"   Will run until Ctrl+C or max {MAX_STEPS} steps\n")
        
        try:
            while step < MAX_STEPS:
                # Check emergency stop IMMEDIATELY at start of loop
                if _emergency_stop:
                    print("\n🛑 Emergency stop detected - Exiting control loop gracefully", flush=True)
                    break  # Exit loop gracefully instead of raising exception
                
                t0 = time.time()
                
                # Check for 'q' key press to stop early (non-blocking)
                import select
                import sys
                if select.select([sys.stdin], [], [], 0.0)[0]:
                    key = sys.stdin.read(1)
                    if key.lower() == 'q':
                        print("\n\n⏹️  'Q' key pressed - Stopping current episode early")
                        break
                
                # Get External ZED image
                ext_img = None
                if zed_ext:
                    ext_img = get_zed_image(zed_ext, zed_ext_runtime)
                    if ext_img is None:
                        continue
            
                # Get Wrist ZED image
                wri_img = None
                if zed_wri:
                    wri_img = get_zed_image(zed_wri, zed_wri_runtime)  # No rotation
                    if wri_img is None:
                        continue
            
                # Check we have both images
                if ext_img is None or wri_img is None:
                    continue
            
                # Capture first frame as preview for this episode
                if step == 0 and episode_preview_image is None:
                    episode_preview_image = ext_img.copy()
                    print(f"   📸 Captured preview image for episode {current_episode}")
                
                # Write frames to video (BGR format for cv2.VideoWriter) - skip if infinite mode
                if not infinite_mode:
                    shoulder_video.write(ext_img)
                    wrist_video.write(wri_img)
            
                # No display during inference (window closed after Enter)
                # Images are still captured for OpenPI inference
            
                # Get current robot state
                joint_pos = robot.get_joint_positions()
                gripper_pos = robot.get_gripper_position()
            
                # Query policy server only when needed (action chunking for efficiency)
                # This is the official OpenPI approach - reduces server load by ~87.5%
                if actions_from_chunk_completed == 0 or actions_from_chunk_completed >= OPEN_LOOP_HORIZON:
                    actions_from_chunk_completed = 0
                
                    # Convert images to RGB and resize to 224x224 (official OpenPI approach)
                    # This reduces network transfer by ~33x (1280x720 -> 224x224)
                    ext_img_rgb = bgr_to_rgb(ext_img)
                    wri_img_rgb = bgr_to_rgb(wri_img)
                    ext_img_resized = resize_with_pad(ext_img_rgb, 224, 224)
                    wri_img_resized = resize_with_pad(wri_img_rgb, 224, 224)
                
                    # Prepare observation for model
                    obs = {
                        "observation/exterior_image_1_left": ext_img_resized,
                        "observation/wrist_image_left": wri_img_resized,
                        "observation/joint_position": joint_pos,
                        "observation/gripper_position": gripper_pos,
                        "prompt": PROMPT,
                    }
                
                    # Get action chunk from OpenPI
                    try:
                        query_start = time.time()
                        out = openpi_client.infer(obs)
                        query_time = time.time() - query_start
                    
                        if "actions" not in out:
                            print(f"✗ No 'actions' key in output: {out.keys()}")
                            break
                    
                        pred_action_chunk = out["actions"]  # shape: (horizon, 8)
                    
                        if len(pred_action_chunk.shape) != 2 or pred_action_chunk.shape[1] != 8:
                            print(f"✗ Unexpected action shape: {pred_action_chunk.shape}, expected (N, 8)")
                            break
                    
                        # if step % 10 == 0:  # Print every 10 steps
                        #     print(f"[Step {step}] Policy query took {query_time*1000:.1f}ms, "
                        #           f"got {pred_action_chunk.shape[0]} actions")
                        
                    except Exception as e:
                        if time.time() - last_print_time > 1.0:
                            print(f"⚠ OpenPI inference error: {e}")
                            last_print_time = time.time()
                        time.sleep(dt)
                        continue
            
                # Execute action from chunk (official OpenPI approach)
                # OpenPI returns: {"actions": (horizon, action_dim)} 
                # PI0: (10, 8), PI05: (15, 8)
                # Each action is [joint_velocity_1...joint_velocity_7, gripper_position]
                try:
                    # Check emergency stop before executing action
                    if _emergency_stop:
                        print("\n🛑 Emergency stop detected - Exiting action execution", flush=True)
                        break  # Exit loop gracefully instead of raising exception
                    
                    action = pred_action_chunk[actions_from_chunk_completed]  # shape: (8,)
                    actions_from_chunk_completed += 1
                
                    # Apply velocity scaling for safety
                    joint_velocities = action[:7] * VELOCITY_SCALE
                    gripper_value = action[7]
                
                    # CRITICAL SAFETY: Prevent hitting table and joint limits
                    # Get current joint positions to monitor
                    current_joints = robot.get_joint_positions()
                
                    # === Joint 1 (shoulder): Safety monitoring only - NO LIMITS ===
                    # J1 positive = arm down (dangerous), J1 negative = arm up (safe)
                    # Monitoring for awareness, but NOT blocking policy actions
                    # if step % 20 == 0:  # Just print status occasionally
                    #     if current_joints[1] > -0.15:
                    #         print(f"  ℹ️  J1={current_joints[1]:+.3f} rad ({np.degrees(current_joints[1]):+.2f}°) - LOW position")
                
                    # NO VELOCITY LIMITS - let policy control freely
                
                    # === Joint 3 (elbow): prevent over-extension ===
                    # Limit: J3 should stay below -2.0 (more negative = safer)
                    if current_joints[3] > -2.0:
                        if joint_velocities[3] > 0:  # Moving toward limit (less negative)
                            joint_velocities[3] = 0.0  # Hard stop
                            if step % 10 == 0:
                                print(f"  🛑 J3={current_joints[3]:+.3f} at limit, blocking extension")
                    elif current_joints[3] > -2.3:
                        if joint_velocities[3] > 0:
                            joint_velocities[3] *= 0.3  # Slow down
                
                    # === Joint 5 (wrist): prevent over-extension forward ===
                    # Limit: J5 should stay below 3.0 (prevent hitting obstacles)
                    if current_joints[5] > 3.0:
                        if joint_velocities[5] > 0:  # Moving forward (increasing)
                            joint_velocities[5] = 0.0  # Hard stop
                            if step % 10 == 0:
                                print(f"  🛑 J5={current_joints[5]:+.3f} at limit, blocking forward")
                    elif current_joints[5] > 2.8:
                        if joint_velocities[5] > 0:
                            joint_velocities[5] *= 0.3  # Slow down
                            if step % 20 == 0:
                                print(f"  ⚠️  J5={current_joints[5]:+.3f} approaching limit")
                
                    # === PI0-specific table height compensation ===
                    # PI0 was trained on DROID official table height, our table is ~2cm higher
                    # Add small downward bias when J1 is still above table level
                    if MODEL_TYPE == "pi0_droid":
                        if current_joints[1] < 0.05 and current_joints[1] > -0.3:  # In approach zone
                            # Add small downward bias (equivalent to ~1cm lower table)
                            joint_velocities[1] += 0.03  # Small positive J1 velocity = downward
                            if step % 20 == 0:
                                print(f"  🔽 PI0 table compensation: J1={current_joints[1]:+.3f} + 0.03 bias")
                
                    # Gripper control - Binary with Hysteresis (prevent oscillation)
                    gripper_value_clipped = np.clip(gripper_value, 0.0, 1.0)
                    
                    # Hysteresis thresholds
                    OPEN_THRESHOLD = 0.6    # Must be > 0.6 to open
                    CLOSE_THRESHOLD = 0.4   # Must be < 0.4 to close
                    
                    # Determine gripper command with hysteresis
                    if last_gripper_cmd is None:
                        # First step: use simple threshold
                        gripper_action = 1.0 if gripper_value_clipped > 0.5 else 0.0
                    else:
                        # Subsequent steps: use hysteresis to prevent oscillation
                        if last_gripper_cmd > 0.5:  # Currently at 1.0
                            # Stay at 1.0 unless clearly told to go to 0.0
                            gripper_action = 0.0 if gripper_value_clipped < CLOSE_THRESHOLD else 1.0
                        else:  # Currently at 0.0
                            # Stay at 0.0 unless clearly told to go to 1.0
                            gripper_action = 1.0 if gripper_value_clipped > OPEN_THRESHOLD else 0.0
                    
                    # Update last command for next iteration
                    last_gripper_cmd = gripper_action
                
                    # Reconstruct action
                    action = np.concatenate([joint_velocities, np.array([gripper_action])])
                
                    # Clip all dimensions to [-1, 1] (official approach)
                    action = np.clip(action, -1, 1)
                
                    # # Print debug info every 10 steps or when querying policy
                    # if step % 10 == 0 or actions_from_chunk_completed == 1:
                    #     print(f"\n{'='*60}")
                    #     print(f"Step {step} - {MODEL_TYPE} (action {actions_from_chunk_completed}/{OPEN_LOOP_HORIZON})")
                    #     print(f"{'='*60}")
                    #     print(f"Current joints: J1={current_joints[1]:+.3f}, J3={current_joints[3]:+.3f}, J5={current_joints[5]:+.3f}")
                    #     print(f"Action (×{VELOCITY_SCALE}): [{action[0]:+.3f}, {action[1]:+.3f}, {action[2]:+.3f}, "
                    #           f"{action[3]:+.3f}, {action[4]:+.3f}, {action[5]:+.3f}, {action[6]:+.3f}, {action[7]:.0f}]")
                    #     print(f"  → Joint vel max: {np.abs(action[:7]).max():.3f}, mean: {np.abs(action[:7]).mean():.3f}")
                    #     print(f"  → Gripper: {action[7]:.0f} ({'OPEN' if action[7] > 0.5 else 'CLOSED'})")
                    #     print(f"{'='*60}\n")
                    
                    # Save combined image when approaching (overwrite each time)
                    if current_joints[5] > 2.6 and action[7] < 0.5 and step % 20 == 0:
                            try:
                                # Resize and stack horizontally
                                h = 480
                                ext_resized = cv2.resize(ext_img, (int(ext_img.shape[1] * h / ext_img.shape[0]), h))
                                wri_resized = cv2.resize(wri_img, (int(wri_img.shape[1] * h / wri_img.shape[0]), h))
                                combined = np.hstack([ext_resized, wri_resized])
                            
                                # Add labels and info
                                cv2.putText(combined, "External", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                                cv2.putText(combined, "Wrist", (ext_resized.shape[1] + 20, 40), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                                cv2.putText(combined, f"Step {step}, J5={current_joints[5]:.2f}", (20, combined.shape[0] - 20), 
                                           cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2)
                            
                                debug_path = Path(__file__).parent / "debug_approach.jpg"
                                cv2.imwrite(str(debug_path), combined)
                                print(f"📸 Updated approach image (J5={current_joints[5]:.3f})")
                            except:
                                pass
                
                    # ===== Execute Action (Official DROID RobotEnv Way) =====
                    # CRITICAL: Use update_command() instead of update_joints()!
                    # update_command() applies proper velocity scaling via IK solver:
                    #   - Converts [-1, 1] velocity to joint_delta via max_joint_delta (0.2 rad)
                    #   - Then converts to joint_position for safe execution
                    # This is why official DROID code doesn't crash into desk!
                
                    # ========================================
                    # DETAILED LOGGING: Model Output → Robot Execution
                    # ========================================
                    # if step % 10 == 0 or actions_from_chunk_completed == 1:
                    #     # Create log message
                    #     log_msg = f"\n{'='*80}\n"
                    #     log_msg += f"STEP {step} | Action {actions_from_chunk_completed}/{OPEN_LOOP_HORIZON} from chunk\n"
                    #     log_msg += f"{'='*80}\n"
                    # 
                    #     # 1. Raw model output (from WebSocket)
                    #     log_msg += f"\n📥 MODEL OUTPUT (raw from policy server):\n"
                    #     log_msg += f"   Shape: {action.shape}\n"
                    #     log_msg += f"   Joint velocities [7]: {action[:7].tolist()}\n"
                    #     log_msg += f"   Gripper position [1]: {action[7]:.6f}\n"
                    #     log_msg += f"   Min: {action.min():.6f}, Max: {action.max():.6f}\n"
                    #     log_msg += f"   Already clipped to [-1, 1]: {(action.min() >= -1) and (action.max() <= 1)}\n"
                    # 
                    #     # 2. After gripper processing (model-specific)
                    #     if MODEL_TYPE == "pi0_droid":
                    #         log_msg += f"\n🔄 AFTER GRIPPER PROCESSING (PI0 - ANALOG):\n"
                    #         log_msg += f"   Gripper: {action[7]:.6f} (continuous 0-1)\n"
                    #     else:
                    #         log_msg += f"\n🔄 AFTER BINARIZATION (PI05):\n"
                    #         log_msg += f"   Gripper: {action[7]:.6f} → {'1.0 (OPEN)' if action[7] > 0.5 else '0.0 (CLOSED)'}\n"
                    # 
                    #     # 3. Current robot state
                    #     log_msg += f"\n🤖 CURRENT ROBOT STATE:\n"
                    #     log_msg += f"   Joint positions [rad]: {[f'{x:+.4f}' for x in current_joints]}\n"
                    #     log_msg += f"   J0: {current_joints[0]:+.4f} rad ({np.degrees(current_joints[0]):+.2f}°)\n"
                    #     log_msg += f"   J1 (shoulder): {current_joints[1]:+.4f} rad ({np.degrees(current_joints[1]):+.2f}°)\n"
                    #     log_msg += f"   J2: {current_joints[2]:+.4f} rad ({np.degrees(current_joints[2]):+.2f}°)\n"
                    #     log_msg += f"   J3 (elbow):    {current_joints[3]:+.4f} rad ({np.degrees(current_joints[3]):+.2f}°)\n"
                    #     log_msg += f"   J4: {current_joints[4]:+.4f} rad ({np.degrees(current_joints[4]):+.2f}°)\n"
                    #     log_msg += f"   J5 (wrist):    {current_joints[5]:+.4f} rad ({np.degrees(current_joints[5]):+.2f}°)\n"
                    #     log_msg += f"   J6: {current_joints[6]:+.4f} rad ({np.degrees(current_joints[6]):+.2f}°)\n"
                    #     log_msg += f"   Gripper state: {gripper_pos:.3f}\n"
                    # 
                    #     # 4. Send to robot
                    #     log_msg += f"\n📤 SENDING TO ROBOT:\n"
                    #     log_msg += f"   action_space: 'joint_velocity'\n"
                    #     log_msg += f"   gripper_action_space: 'position'\n"
                    #     log_msg += f"   blocking: False\n"
                    # 
                    #     # Print to console
                    #     print(log_msg, end='')
                
                    action_dict = robot.update_command(
                        command=action.tolist(),
                        action_space="joint_velocity",      # DROID uses joint velocity
                        gripper_action_space="position",    # Gripper uses position (0=closed, 1=open)
                        blocking=False,  # Non-blocking for smooth 15Hz real-time control
                        use_true_velocity=True,  # Always use true velocity control
                        model_type=MODEL_TYPE  # Pass model type for velocity scaling
                    )
                
                    # if step % 10 == 0 or actions_from_chunk_completed == 1:
                    #     if action_dict:
                    #         # 5. Robot's computed target (after IK solver conversion)
                    #         log_msg = f"\n✅ ROBOT COMPUTED TARGET:\n"
                    #         if "joint_position" in action_dict:
                    #             target_joints = action_dict['joint_position']
                    #             log_msg += f"   Target joint_position [rad]: {[f'{x:+.4f}' for x in target_joints]}\n"
                    #         
                    #             # Calculate deltas
                    #             deltas = [target_joints[i] - current_joints[i] for i in range(7)]
                    #             log_msg += f"\n📐 COMPUTED DELTAS (target - current):\n"
                    #             log_msg += f"   Joint deltas [rad]: {[f'{x:+.4f}' for x in deltas]}\n"
                    #             log_msg += f"   J0 delta: {deltas[0]:+.4f} rad ({np.degrees(deltas[0]):+.2f}°)\n"
                    #             log_msg += f"   J1 delta: {deltas[1]:+.4f} rad ({np.degrees(deltas[1]):+.2f}°)\n"
                    #             log_msg += f"   J2 delta: {deltas[2]:+.4f} rad ({np.degrees(deltas[2]):+.2f}°)\n"
                    #             log_msg += f"   J3 delta: {deltas[3]:+.4f} rad ({np.degrees(deltas[3]):+.2f}°)\n"
                    #             log_msg += f"   J4 delta: {deltas[4]:+.4f} rad ({np.degrees(deltas[4]):+.2f}°)\n"
                    #             log_msg += f"   J5 delta: {deltas[5]:+.4f} rad ({np.degrees(deltas[5]):+.2f}°)\n"
                    #             log_msg += f"   J6 delta: {deltas[6]:+.4f} rad ({np.degrees(deltas[6]):+.2f}°)\n"
                    #         
                    #             # Velocity control: velocity → rad/s (NUC multiplies by max_vel) → delta over time
                    #             log_msg += f"\n🔬 VELOCITY ANALYSIS:\n"
                    #             log_msg += f"   Normalized velocities: {[f'{x:+.4f}' for x in action[:7]]}\n"
                    #             log_msg += f"   Actual deltas (from robot): {[f'{x:+.4f}' for x in deltas]}\n"
                    #         
                    #             # Specific joint analysis
                    #             log_msg += f"\n🎯 CRITICAL JOINT MOVEMENTS:\n"
                    #             log_msg += f"   J1 (shoulder):\n"
                    #             log_msg += f"     Current:   {current_joints[1]:+.4f} rad ({np.degrees(current_joints[1]):+.2f}°)\n"
                    #             log_msg += f"     Velocity:  {action[1]:+.4f} (normalized)\n"
                    #             log_msg += f"     Delta:     {deltas[1]:+.4f} rad ({np.degrees(deltas[1]):+.2f}°)\n"
                    #             log_msg += f"     Target:    {target_joints[1]:+.4f} rad ({np.degrees(target_joints[1]):+.2f}°)\n"
                    #             log_msg += f"     Direction: {'⬇ DOWNWARD (toward table!)' if deltas[1] > 0 else '⬆ UPWARD (safe)'}\n"
                    #     
                    #         if "gripper_position" in action_dict:
                    #             log_msg += f"\n🦾 GRIPPER:\n"
                    #             log_msg += f"   Target: {action_dict['gripper_position']:.3f}\n"
                    #     
                    #         log_msg += f"{'='*80}\n\n"
                    #     
                    #         # Print to console
                    #         print(log_msg, end='')
                
                    # ===== End Gripper Control =====
                
                except Exception as e:
                    print(f"✗ Action execution error: {e}")
                    import traceback
                    traceback.print_exc()
                    break
            
                step += 1
            
                # Progress bar display
                progress = step / MAX_STEPS * 100
                bar_length = 40
                filled = int(bar_length * step / MAX_STEPS)
                bar = '█' * filled + '-' * (bar_length - filled)
                print(f'\rStep [{bar}] {step}/{MAX_STEPS} ({progress:.1f}%)', end='', flush=True)
            
                # Check if we've reached max steps
                if step >= MAX_STEPS:
                    print(f"\n✓ Reached maximum steps ({MAX_STEPS}). Stopping...")
                    break
            
                # Sleep to match DROID control frequency (official approach)
                # This ensures consistent timing even if inference is slow
                elapsed_time = time.time() - t0
                if elapsed_time < dt:
                    time.sleep(dt - elapsed_time)
            
                # Monitor loop time
                if elapsed_time > dt * 1.5 and step % 10 == 0:
                    print(f"⚠ Loop running slow: {elapsed_time*1000:.1f}ms (target: {dt*1000:.1f}ms)")
                    
        except KeyboardInterrupt:
            print("\n\n🛑 EMERGENCY STOP - Ctrl+C pressed")
            print("   Stopping robot immediately...")
            
            # PRIORITY 1: Stop robot immediately by sending zero velocity
            try:
                zero_velocity = [0.0] * 7 + [robot.get_gripper_position()]
                robot.update_command(
                    command=zero_velocity,
                    action_space="joint_velocity",
                    gripper_action_space="position",
                    blocking=False
                )
                time.sleep(0.1)  # Give it a moment to stop
                print("   ✓ Robot stopped")
            except Exception as e:
                print(f"   ⚠ Stop error: {e}")
            
            # PRIORITY 1.5: Reset OpenPI client to stop inference
            try:
                openpi_client.reset()
                print("   ✓ OpenPI client reset")
            except Exception as e:
                print(f"   ⚠ OpenPI reset error: {e}")
            
            # PRIORITY 2: Save videos (skip if infinite mode)
            if not infinite_mode:
                try:
                    shoulder_video.release()
                    wrist_video.release()
                    print(f"   ✓ Videos saved to: {session_dir}")
                except Exception as e:
                    print(f"   ⚠ Video save error: {e}")
                
                # Update instruction file
                try:
                    with open(instruction_file, "r") as f:
                        content = f.read()
                    content = content.replace("Steps: TBD", f"Steps: {step}")
                    content = content.replace("Success: TBD", f"Success: {success_str}")
                    with open(instruction_file, "w") as f:
                        f.write(content)
                except:
                    pass
            
            # PRIORITY 3: Return to home position
            try:
                print("\n   Moving robot to home position...")
                home_joints = np.array([0.0, -0.5, 0.0, -2.40, 0.0, 1.90, 0.0])
                robot.update_joints(command=home_joints.tolist(), velocity=False, blocking=True)
                robot.update_gripper(command=0.0, velocity=False, blocking=True)
                print("   ✓ Robot returned to home position")
            except Exception as e:
                print(f"   ⚠ Reset error: {e}")
            
            # Exit immediately
            cleanup_cameras()
            print("\n✓ Program terminated by user")
            sys.exit(0)  # Force exit to prevent any continued execution
            
        except Exception as e:
            print(f"\n✗ Error: {e}")
            import traceback
            traceback.print_exc()
        finally:
            # Normal completion (not emergency stop)
            print("\n📹 Episode Complete!")
            print(f"   Total steps: {step}")
            
            # Close video files for this episode - CRITICAL for re-recording (skip if infinite mode)
            if not infinite_mode:
                try:
                    if 'shoulder_video' in locals() and shoulder_video is not None:
                        shoulder_video.release()
                    if 'wrist_video' in locals() and wrist_video is not None:
                        wrist_video.release()
                    print(f"   ✓ Videos saved to: {session_dir}")
                except Exception as e:
                    print(f"   ⚠ Video save error: {e}")
                
                # Update instruction file with actual step count
                try:
                    with open(instruction_file, "r") as f:
                        content = f.read()
                    modified = False
                    if "Steps: TBD" in content:
                        content = content.replace("Steps: TBD", f"Steps: {step}")
                        modified = True
                    if "Success: TBD" in content:
                        content = content.replace("Success: TBD", "Success: No")  # Interrupted, mark as failed
                        modified = True
                    if modified:
                        with open(instruction_file, "w") as f:
                            f.write(content)
                        print(f"   ✓ Instruction file updated")
                except Exception as e:
                    print(f"   ⚠ File save error: {e}")
        
        # Reset robot to home position after each episode
        print(f"\n🔄 Resetting robot...")
        try:
            # Move to home position
            home_joints = np.array([
                0.0, -0.5, 0.0, -2.40, 
                0.0, 1.90, 0.0
            ])
            
            current_joints = robot.get_joint_positions()
            print(f"  Moving to home position...")
            print(f"    Current: {np.round(current_joints, 3).tolist()}")
            print(f"    Target:  {np.round(home_joints, 3).tolist()}")
            
            robot.update_joints(
                command=home_joints.tolist(),
                velocity=False,
                blocking=True
            )
            
            # Open gripper
            robot.update_gripper(command=0.0, velocity=False, blocking=True)
            print("  ✓ Robot reset complete")
            
            # Check emergency stop before asking for input
            if _emergency_stop:
                print("\n🛑 Emergency stop detected - Exiting without prompting")
                break
            
            # In infinite mode (LOOP=0), just exit after reset - no re-recording option
            if infinite_mode:
                print("\n✓ Episode complete (infinite mode)")
                print("   Exiting program...")
                break
            
            # Ask user what to do next (only for LOOP > 0)
            print(f"\n⏸️  Episode {current_episode} complete (Progress: {completed_episodes}/{LOOP}). Options:")
            if completed_episodes < num_iterations - 1:
                # Not the last episode yet
                print(f"   Press ENTER to continue to next episode")
                print(f"   Press 'R' + ENTER to re-record this episode")
                print(f"   Press Ctrl+C to exit program")
            else:
                # This would be the last episode if accepted
                print(f"   Press ENTER to finish (this will complete {LOOP}/{LOOP})")
                print(f"   Press 'R' + ENTER to re-record this episode")
                print(f"   Press Ctrl+C to exit program")
            
            try:
                user_input = input("   Your choice: ").strip().lower()
            except (KeyboardInterrupt, EOFError):
                # User pressed Ctrl+C or Ctrl+D during input
                print("\n\n🛑 User interrupted - Exiting...")
                _emergency_stop = True
                break
            
            if user_input == 'r':
                print(f"\n🔄 Re-recording episode {current_episode}...")
                # Delete the current episode's videos (session_dir is guaranteed to exist here since infinite_mode=False)
                import shutil
                import time as time_module
                
                # Give a moment for file handles to be fully released
                time_module.sleep(0.5)
                
                try:
                    if 'session_dir' in locals():
                        print(f"   Deleting: {session_dir}")
                        if session_dir.exists():
                            shutil.rmtree(session_dir)
                            print(f"   ✓ Deleted previous recording")
                        else:
                            print(f"   ⚠ Directory not found: {session_dir}")
                    else:
                        print(f"   ⚠ No session directory to delete (infinite mode?)")
                except PermissionError as e:
                    print(f"   ✗ Permission denied - files may still be in use")
                    print(f"   Waiting and retrying...")
                    time_module.sleep(1.0)
                    try:
                        shutil.rmtree(session_dir)
                        print(f"   ✓ Deleted on retry")
                    except Exception as e2:
                        print(f"   ✗ Still failed: {e2}")
                except Exception as e:
                    print(f"   ✗ Could not delete: {e}")
                    import traceback
                    traceback.print_exc()
                
                # Don't increment anything - will repeat this episode
            else:
                # User pressed Enter - accept this episode and move to next
                
                # Excel logging (if enabled)
                if ENABLE_EXCEL_LOGGING and 'session_dir' in locals():
                    if not PANDAS_AVAILABLE:
                        print("\n⚠ Excel logging enabled but pandas not installed. Skipping log.")
                        print("   Install with: pip install pandas openpyxl")
                    else:
                        try:
                            # Auto-generate model name based on MODEL_TYPE
                            if session_model_name is None:
                                if "pi0_fast" in MODEL_TYPE.lower():
                                    session_model_name = "Pi0-fast-velocity"
                                elif "pi0" in MODEL_TYPE.lower() and "pi05" not in MODEL_TYPE.lower():
                                    session_model_name = "Pi0-velocity"
                                elif "pi05" in MODEL_TYPE.lower():
                                    session_model_name = "Pi0.5-velocity"
                                else:
                                    session_model_name = "Unknown-velocity"
                                print(f"\n📝 Auto-detected model: {session_model_name}")
                            
                            # Ask for success status with retry
                            success_str = None
                            while success_str is None:
                                success_input = input("   Success? (y/n): ").strip().lower()
                                
                                if success_input == 'y':
                                    success_str = 'Yes'
                                elif success_input == 'n':
                                    success_str = 'No'
                                else:
                                    print("   ⊘ Invalid input. Please enter 'y' or 'n'")
                            
                            # Prepare log data (Preview will be added as first column)
                            log_data = {
                                'Task': PROMPT,
                                'Model': session_model_name,
                                'Position': POSITION_VARIANT if ENABLE_POSITION_VARIANT else None,
                                'Episode': current_episode,
                                'Success': success_str,
                                'Video_Path': str((session_dir / 'shoulder_view.mp4').relative_to(Path(__file__).parent)),
                                'Steps': step
                            }
                            
                            # Update instruction.txt with Success value
                            try:
                                with open(instruction_file, "r") as f:
                                    content = f.read()
                                # Replace both TBD and No (in case finally block already wrote "No")
                                content = content.replace("Success: TBD", f"Success: {success_str}")
                                content = content.replace("Success: No", f"Success: {success_str}")
                                with open(instruction_file, "w") as f:
                                    f.write(content)
                            except:
                                pass
                            
                            # Excel file path in vid directory
                            excel_path = vid_base_dir / 'episode_log.xlsx'
                            
                            # Save to Excel with embedded preview images
                            try:
                                from openpyxl import load_workbook
                                from openpyxl.drawing.image import Image as XLImage
                                from openpyxl import Workbook
                                import io
                                
                                # Define expected column order based on ENABLE_POSITION_VARIANT
                                if ENABLE_POSITION_VARIANT:
                                    expected_header = ['Preview', 'Task', 'Model', 'Position', 'Episode', 'Success', 'Video_Path', 'Steps']
                                else:
                                    expected_header = ['Preview', 'Task', 'Model', 'Episode', 'Success', 'Video_Path', 'Steps']
                                
                                if excel_path.exists():
                                    # Load existing workbook and verify format
                                    wb = load_workbook(excel_path)
                                    ws = wb.active
                                    
                                    # Check existing header
                                    existing_header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
                                    
                                    if existing_header != expected_header:
                                        print(f"\n❌ ERROR: Excel format mismatch!")
                                        print(f"   Expected: {expected_header}")
                                        print(f"   Found:    {existing_header}")
                                        print(f"\n   ENABLE_POSITION_VARIANT = {ENABLE_POSITION_VARIANT}")
                                        print(f"   Please check your Excel file format or change ENABLE_POSITION_VARIANT setting.")
                                        print(f"   Program will exit to prevent data corruption.")
                                        import sys
                                        cleanup_cameras()
                                        sys.exit(1)
                                    
                                    # Append new row data (skip Preview column A)
                                    if ENABLE_POSITION_VARIANT:
                                        new_row = [
                                            "",  # Preview column (will add image later)
                                            log_data['Task'],
                                            log_data['Model'],
                                            log_data['Position'],
                                            log_data['Episode'],
                                            log_data['Success'],
                                            log_data['Video_Path'],
                                            log_data['Steps']
                                        ]
                                    else:
                                        new_row = [
                                            "",  # Preview column (will add image later)
                                            log_data['Task'],
                                            log_data['Model'],
                                            log_data['Episode'],
                                            log_data['Success'],
                                            log_data['Video_Path'],
                                            log_data['Steps']
                                        ]
                                    ws.append(new_row)
                                    row_num = ws.max_row
                                    
                                else:
                                    # Create new workbook
                                    wb = Workbook()
                                    ws = wb.active
                                    
                                    # Write header
                                    ws.append(expected_header)
                                    
                                    # Write first data row
                                    if ENABLE_POSITION_VARIANT:
                                        ws.append([
                                            "",
                                            log_data['Task'],
                                            log_data['Model'],
                                            log_data['Position'],
                                            log_data['Episode'],
                                            log_data['Success'],
                                            log_data['Video_Path'],
                                            log_data['Steps']
                                        ])
                                    else:
                                        ws.append([
                                            "",
                                            log_data['Task'],
                                            log_data['Model'],
                                            log_data['Episode'],
                                            log_data['Success'],
                                            log_data['Video_Path'],
                                            log_data['Steps']
                                        ])
                                    row_num = 2  # First data row
                                    
                                    # Set column width for preview
                                    ws.column_dimensions['A'].width = 30
                                
                                # Add preview image for current episode
                                if 'episode_preview_image' in locals() and episode_preview_image is not None:
                                    # Resize to preview size
                                    preview_img = cv2.resize(episode_preview_image, (320, 180))
                                    # Convert BGR to RGB
                                    preview_img_rgb = cv2.cvtColor(preview_img, cv2.COLOR_BGR2RGB)
                                    # Save to memory buffer
                                    from PIL import Image as PILImage
                                    pil_img = PILImage.fromarray(preview_img_rgb)
                                    img_buffer = io.BytesIO()
                                    pil_img.save(img_buffer, format='PNG')
                                    img_buffer.seek(0)
                                    
                                    # Add image to current row
                                    ws.row_dimensions[row_num].height = 135
                                    img = XLImage(img_buffer)
                                    img.width = 320
                                    img.height = 180
                                    img.anchor = f'A{row_num}'
                                    ws.add_image(img)
                                
                                wb.save(excel_path)
                                print(f"   ✓ Logged to: {excel_path}")
                            except ImportError:
                                # Fallback: save without images if openpyxl not available
                                import pandas as pd
                                # Add Preview column for consistency
                                fallback_data = {'Preview': '', **log_data}
                                if excel_path.exists():
                                    df = pd.read_excel(excel_path)
                                    df = pd.concat([df, pd.DataFrame([fallback_data])], ignore_index=True)
                                else:
                                    df = pd.DataFrame([fallback_data])
                                df.to_excel(excel_path, index=False)
                                print(f"   ⚠ openpyxl not available - saved without preview image")
                                print(f"   ✓ Logged to: {excel_path}")
                            except Exception as e:
                                print(f"   ⚠ Image embedding failed: {e}")
                                import traceback
                                traceback.print_exc()
                                # Still save the data even if image embedding fails
                                import pandas as pd
                                # Add Preview column for consistency
                                fallback_data = {'Preview': '', **log_data}
                                if excel_path.exists():
                                    df = pd.read_excel(excel_path)
                                    df = pd.concat([df, pd.DataFrame([fallback_data])], ignore_index=True)
                                else:
                                    df = pd.DataFrame([fallback_data])
                                df.to_excel(excel_path, index=False)
                                print(f"   ✓ Logged to: {excel_path} (data only)")
                        except KeyboardInterrupt:
                            print("\n   ⊘ Logging cancelled")
                        except Exception as e:
                            print(f"\n   ✗ Logging failed: {e}")
                
                completed_episodes += 1
                current_episode += 1
                print(f"   ✓ Episode accepted ({completed_episodes}/{LOOP} completed)")
            
        except Exception as e:
            print(f"  ✗ Reset failed: {e}")
            break
    
    # All episodes complete - final cleanup
    print("\n" + "="*70)
    
    # Check if exited due to emergency stop
    if _emergency_stop:
        print("🛑 Exited due to emergency stop")
        print("   Performing cleanup...")
    elif LOOP > 1:
        print(f"🎉 All {LOOP} episodes completed!")
        print(f"   Videos saved in: {task_dir}")
    elif LOOP == 1:
        print("🎉 Recording complete!")
        if 'session_dir' in locals():
            print(f"   Video saved in: {session_dir}")
    else:
        print("🎉 Session complete (no recording in LOOP=0 mode)")
    print("="*70)
    
    # Final cleanup
    print("\n[Final Cleanup]")
    try:
        # PRIORITY 1: Reset robot to safe position
        print("  [1/2] Moving robot to home position...")
        home_joints = np.array([
            0.0, -0.5, 0.0, -2.40, 
            0.0, 1.90, 0.0
        ])
        
        current_joints = robot.get_joint_positions()
        print(f"    Current: {np.round(current_joints, 3).tolist()}")
        print(f"    Target:  {np.round(home_joints, 3).tolist()}")
        
        robot.update_joints(
            command=home_joints.tolist(),
            velocity=False,
            blocking=True
        )
        
        robot.update_gripper(command=0.0, velocity=False, blocking=True)
        print("  ✓ Robot at home position")
        
    except Exception as e:
        print(f"  ✗ Reset failed: {e}")
    
    try:
        # PRIORITY 2: Close cameras
        print("  [2/2] Closing cameras...")
        cv2.destroyAllWindows()
        cleanup_cameras()
        print("  ✓ Cameras closed")
        
    except Exception as e:
        print(f"  ⚠ Cleanup error: {e}")
    
    print("\n✓ Program ended successfully")


if __name__ == "__main__":
    main()

"""
Quick Reference - Model Switching:

1. To use PI0 model:
   - Change: MODEL_TYPE = "pi0_droid"
   - Server: python scripts/serve_policy.py --port 5555 policy:checkpoint \
             --policy.config=pi0_droid \
             --policy.dir=gs://openpi-assets/checkpoints/pi0_droid

2. To use PI05 model:
   - Change: MODEL_TYPE = "pi05_droid"
   - Server: python scripts/serve_policy.py --port 5555 policy:checkpoint \
             --policy.config=pi05_droid \
             --policy.dir=gs://openpi-assets/checkpoints/pi05_droid

Model Differences:
  PI0:  action_horizon=10, optimized for speed
  PI05: action_horizon=16, improved accuracy

Both models use same robot interface and control parameters.
"""
